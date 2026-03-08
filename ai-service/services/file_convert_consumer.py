"""Consumer callback for file_convert_queue.

Handles file conversion logic (Phase 2c):
- PDF  → text extraction (PyMuPDF), fallback to page images if text insufficient
- DOCX → text extraction (python-docx)
- XLSX → Markdown table extraction (openpyxl)
- PPTX → page images (LibreOffice headless → PDF → PyMuPDF)
"""

from __future__ import annotations

import asyncio
import io
import json
import logging
import subprocess
import tempfile
from pathlib import Path

import fitz  # PyMuPDF
import openpyxl
from aio_pika.abc import AbstractIncomingMessage
from docx import Document as DocxDocument
from PIL import Image

from services import minio_client, redis_client

logger = logging.getLogger(__name__)

# ── Constants ──
TEXT_MAX_BYTES = 50 * 1024  # 50 KB text truncation limit
TEXT_TRUNCATION_MARKER = "\n[...text truncated]"
IMAGE_MAX_WIDTH = 1568
IMAGE_JPEG_QUALITY = 80
PDF_MIN_CHARS_PER_PAGE = 100  # threshold to decide text vs image route
SUMMARY_TEXT_THRESHOLD = 10 * 1024  # 10 KB — generate summary if text exceeds this


# ── Image helpers ──

def _compress_image(data: bytes) -> bytes:
    """Compress an image to JPEG, max width 1568px, quality 80."""
    img = Image.open(io.BytesIO(data))
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    if img.width > IMAGE_MAX_WIDTH:
        ratio = IMAGE_MAX_WIDTH / img.width
        img = img.resize((IMAGE_MAX_WIDTH, int(img.height * ratio)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=IMAGE_JPEG_QUALITY)
    return buf.getvalue()


def _render_pdf_page_to_jpeg(page: fitz.Page) -> bytes:
    """Render a single PDF page to compressed JPEG bytes."""
    # Render at 2x for clarity, then compress
    pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
    return _compress_image(pix.tobytes("png"))


# ── Text helpers ──

def _truncate_text(text: str) -> str:
    """Truncate text to TEXT_MAX_BYTES, appending a marker if truncated."""
    encoded = text.encode("utf-8")
    if len(encoded) <= TEXT_MAX_BYTES:
        return text
    # Truncate at byte boundary then decode safely
    truncated = encoded[: TEXT_MAX_BYTES - len(TEXT_TRUNCATION_MARKER.encode("utf-8"))]
    return truncated.decode("utf-8", errors="ignore") + TEXT_TRUNCATION_MARKER


# ── Per-type converters ──
# Each returns (text | None, page_images | None)
# text: extracted text string (to be stored as _text.txt)
# page_images: list of (page_name, jpeg_bytes) tuples (to be stored under _pages/)

def _convert_pdf(file_bytes: bytes) -> tuple[str | None, list[tuple[str, bytes]] | None]:
    """Extract text from PDF; fallback to page images if text is insufficient."""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    try:
        # Try text extraction first
        pages_text: list[str] = []
        total_chars = 0
        for page in doc:
            text = page.get_text()
            pages_text.append(text)
            total_chars += len(text)

        avg_chars = total_chars / max(len(doc), 1)
        if avg_chars >= PDF_MIN_CHARS_PER_PAGE:
            # Sufficient text — use text route
            full_text = "\n\n".join(pages_text)
            return _truncate_text(full_text), None

        # Insufficient text (scanned PDF) — render pages as images
        images: list[tuple[str, bytes]] = []
        for i, page in enumerate(doc):
            jpeg = _render_pdf_page_to_jpeg(page)
            images.append((f"page_{i + 1}.jpg", jpeg))
        return None, images
    finally:
        doc.close()


def _convert_docx(file_bytes: bytes) -> tuple[str | None, list[tuple[str, bytes]] | None]:
    """Extract text from DOCX using python-docx."""
    doc = DocxDocument(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    full_text = "\n".join(paragraphs)
    return _truncate_text(full_text), None


def _convert_xlsx(file_bytes: bytes) -> tuple[str | None, list[tuple[str, bytes]] | None]:
    """Extract XLSX content as Markdown tables (one table per sheet)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parts.append(f"## {sheet_name}\n")
        # Header
        header = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        # Data rows
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("| " + " | ".join(cells) + " |")
        parts.append("")  # blank line between sheets
    wb.close()
    full_text = "\n".join(parts)
    return _truncate_text(full_text), None


def _find_libreoffice() -> str:
    """Return the LibreOffice CLI command available on this system."""
    import shutil
    for cmd in ("libreoffice", "soffice"):
        if shutil.which(cmd):
            return cmd
    raise FileNotFoundError("LibreOffice is not installed or not in PATH")


def _convert_pptx(file_bytes: bytes) -> tuple[str | None, list[tuple[str, bytes]] | None]:
    """Convert PPTX to page images via LibreOffice headless → PDF → PyMuPDF."""
    lo_cmd = _find_libreoffice()
    with tempfile.TemporaryDirectory() as tmpdir:
        pptx_path = Path(tmpdir) / "input.pptx"
        pptx_path.write_bytes(file_bytes)

        # LibreOffice headless export to PDF
        subprocess.run(
            [
                lo_cmd, "--headless", "--convert-to", "pdf",
                "--outdir", tmpdir, str(pptx_path),
            ],
            check=True,
            timeout=120,
            capture_output=True,
        )

        pdf_path = Path(tmpdir) / "input.pdf"
        if not pdf_path.exists():
            raise RuntimeError("LibreOffice failed to produce PDF from PPTX")

        # Render PDF pages to images
        doc = fitz.open(str(pdf_path))
        try:
            images: list[tuple[str, bytes]] = []
            for i, page in enumerate(doc):
                jpeg = _render_pdf_page_to_jpeg(page)
                images.append((f"page_{i + 1}.jpg", jpeg))
            return None, images
        finally:
            doc.close()


# ── Summary generation ──

async def _generate_summary(text: str, minio_path: str) -> None:
    """Generate a summary for large text files and write to MinIO.

    Called asynchronously after the main conversion is done.
    Failure does not affect the main flow.
    """
    try:
        from services.llm import generate_title  # avoid circular at top-level
        from config import settings

        entry = settings.utility_model
        provider = entry["provider"]
        model = entry["model"]

        system_prompt = (
            "You are a document summarization assistant. Generate a concise summary of the "
            "following document content in no more than 500 words. Preserve the core points "
            "and key data. Return only the summary text — no extra commentary. "
            "Write the summary in the same language as the document."
        )

        # Use the same LLM infrastructure as title generation
        if provider == "claude":
            import anthropic
            client = anthropic.AsyncAnthropic(api_key=settings.claude_api_key)
            response = await client.messages.create(
                model=model,
                max_tokens=1024,
                system=system_prompt,
                messages=[{"role": "user", "content": text}],
            )
            summary = response.content[0].text.strip()
        elif provider == "gemini":
            from google import genai
            client = genai.Client(api_key=settings.gemini_api_key)
            response = await client.aio.models.generate_content(
                model=model,
                contents=text,
                config=genai.types.GenerateContentConfig(
                    system_instruction=system_prompt,
                    max_output_tokens=1024,
                ),
            )
            summary = response.text.strip() if response.text else ""
        else:
            from openai import AsyncOpenAI
            creds = {
                "deepseek": (settings.deepseek_api_key, settings.deepseek_base_url),
                "openai": (settings.openai_api_key, settings.openai_base_url),
            }
            api_key, base_url = creds[provider]
            client = AsyncOpenAI(api_key=api_key, base_url=base_url)
            response = await client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": text},
                ],
                max_tokens=1024,
            )
            summary = response.choices[0].message.content.strip()

        if summary:
            summary_path = f"{minio_path}_summary.txt"
            minio_client.write_text_file(summary_path, summary)
            logger.info("summary written: %s", summary_path)
    except Exception:
        logger.exception("summary generation failed (non-fatal): %s", minio_path)


# ── Router ──

_CONVERTERS = {
    "application/pdf": _convert_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": _convert_docx,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": _convert_xlsx,
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": _convert_pptx,
}


# ── Main callback ──

async def file_convert_callback(message: AbstractIncomingMessage) -> None:
    """Process a file conversion task from RabbitMQ.

    Message body: {"file_id": "...", "minio_path": "...", "content_type": "..."}
    """
    file_id = "unknown"
    try:
        payload = json.loads(message.body.decode())
        file_id = payload.get("file_id", "unknown")
        content_type = payload.get("content_type", "")
        minio_path = payload.get("minio_path", "")

        logger.info(
            "received file_convert task: file_id=%s content_type=%s path=%s",
            file_id, content_type, minio_path,
        )

        rd = redis_client.get_client()

        # 1. Idempotency check
        exists = await rd.get(f"file:wait_to_process:{file_id}")
        if not exists:
            logger.info("file_convert skipped (idempotent): file_id=%s", file_id)
            await message.ack()
            return

        # 2. Find converter
        converter = _CONVERTERS.get(content_type)
        if converter is None:
            logger.warning("no converter for content_type=%s, skipping file_id=%s", content_type, file_id)
            await rd.delete(f"file:processing:{file_id}", f"file:wait_to_process:{file_id}")
            await message.ack()
            return

        # 3. Download file from MinIO
        file_bytes = minio_client.read_file(minio_path)

        # 4. Run conversion (CPU-bound, run in thread pool)
        loop = asyncio.get_running_loop()
        text, page_images = await loop.run_in_executor(None, converter, file_bytes)

        # 5. Write results to MinIO
        if text is not None:
            text_path = f"{minio_path}_text.txt"
            minio_client.write_text_file(text_path, text)
            logger.info("text written: %s (%d bytes)", text_path, len(text.encode("utf-8")))

        if page_images is not None:
            for page_name, jpeg_bytes in page_images:
                page_path = f"{minio_path}_pages/{page_name}"
                minio_client.write_file(page_path, jpeg_bytes, content_type="image/jpeg")
            logger.info("pages written: %s_pages/ (%d pages)", minio_path, len(page_images))

        # 6. Release locks → ACK (release blocking lock first)
        await rd.delete(f"file:processing:{file_id}", f"file:wait_to_process:{file_id}")
        await message.ack()
        logger.info("file_convert task acked: file_id=%s", file_id)

        # 7. Async summary generation (after ACK, non-blocking)
        if text is not None and len(text.encode("utf-8")) > SUMMARY_TEXT_THRESHOLD:
            asyncio.create_task(_generate_summary(text, minio_path))

    except Exception:
        logger.exception("file_convert task failed, rejecting to DLQ: file_id=%s", file_id)
        try:
            rd = redis_client.get_client()
            await rd.delete(f"file:processing:{file_id}", f"file:wait_to_process:{file_id}")
        except Exception:
            logger.exception("failed to clean up Redis keys for file_id=%s", file_id)
        await message.reject(requeue=False)
